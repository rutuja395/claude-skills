#!/usr/bin/env python3
"""
Generate spending analysis Excel report from transaction CSV.
Works for any currency, any bank, any country.
"""
import argparse, re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference

# ── Styles ───────────────────────────────────────────────────────────────────
NAVY="FF1F3864"; TEAL="FF0097A7"; ORANGE="FFFF6F00"; GOLD="FFFFC107"
WHITE="FFFFFFFF"; LGRAY="FFF5F5F5"; RED="FFD32F2F"; GREEN="FF2E7D32"; PURPLE="FF6A1B9A"

def fill(h): return PatternFill("solid", start_color=h, end_color=h)
def hf(size=11,bold=True,color=WHITE): return Font(name='Arial',size=size,bold=bold,color=color)
def bf(size=10,bold=False,color='FF000000'): return Font(name='Arial',size=size,bold=bold,color=color)
def ctr(): return Alignment(horizontal='center',vertical='center',wrap_text=True)
def lft(): return Alignment(horizontal='left',vertical='center',wrap_text=True)
def rgt(): return Alignment(horizontal='right',vertical='center')
def bdr():
    s=Side(style='thin',color='FFCCCCCC')
    return Border(left=s,right=s,top=s,bottom=s)

CAT_COLORS = {
    'Food & Dining':'FFE3F2FD','Groceries':'FFE8F5E9','Shopping':'FFFFF3E0',
    'Travel & Transport':'FFF3E5F5','Subscriptions':'FFFCE4EC','Health & Fitness':'FFE0F2F1',
    'Entertainment':'FFFFF8E1','Electronics':'FFE8EAF6','Luxury':'FFF8F0FF',
    'Rent & Housing':'FFEFEBE9','Utilities':'FFE0F7FA','Personal Care':'FFFFF9C4',
    'Telecom':'FFF1F8E9','Fuel':'FFFFF3E0','Education':'FFE8EAF6',
    'Insurance':'FFE0F7FA','Investments':'FFE8F5E9','Transfers':'FFFAFAFA',
    'Miscellaneous':'FFFAFAFA',
}
NON_SPEND = {'Investments','Transfers'}

# ── Categorisation (global keywords) ─────────────────────────────────────────
CATEGORY_RULES = [
    ('Investments', [r'ZERODHA',r'GROWW',r'KUVERA',r'VANGUARD',r'FIDELITY',r'SCHWAB',
                     r'ETRADE',r'ROBINHOOD',r'TRADING\s*212',r'FREETRADE',r'NUTMEG',
                     r'WEALTHSIMPLE',r'STAKE',r'NACH.*MF',r'MUTUAL\s*FUND',r'SIP\b',
                     r'BROKER',r'SHARES',r'STOCKS',r'ETF\b',r'PENSION',r'ISA\b']),
    ('Rent & Housing', [r'RENT\b',r'MORTGAGE',r'LANDLORD',r'LETTING',r'LEASE',
                        r'HOUSE\s*PAYMENT',r'PROPERTY',r'ACCOMMODATION',r'FLAT\b',
                        r'HOUSING',r'ESTATE\s*AGENT',r'NOBROKER',r'NESTAWAY']),
    ('Transfers', [r'TRANSFER\s*TO\s*OWN',r'OWN\s*ACCOUNT',r'SELF\s*TRANSFER',
                   r'SENTIMPS',r'NEFT\b',r'RTGS\b',r'IMPS\b']),
    ('Food & Dining', [r'ZOMATO',r'SWIGGY',r'BUNDLTECHNOLOG',r'BLINKIT',
                       r'DELIVEROO',r'UBER\s*EATS',
                       r'DOORDASH',r'JUST\s*EAT',r'GRUBHUB',r'MENULOG',r'FOODPANDA',
                       r'MCDONALD',r'KFC\b',r'BURGER\s*KING',r'SUBWAY\b',r'PIZZA\s*HUT',
                       r'DOMINO',r'STARBUCKS',r'COSTA\b',r'PRET\b',r'GREGGS',
                       r'CAFE\b',r'RESTAURANT',r'BISTRO',r'DINER\b',r'EATERY',
                       r'CANTEEN',r'FOOD\s*COURT',r'TAKEAWAY',r'TAKEOUT',r'DINING',
                       r'BAKERY',r'COFFEE\s*SHOP',r'NANDO',r'WAGAMAMA',r'ZIZZI',
                       r'CHAAYOS',r'CHAI\s*POINT',r'BARISTA\b',r'SOCIAL\b']),
    ('Groceries', [r'TESCO\b',r'SAINSBURY',r'ASDA\b',r'MORRISONS',r'WAITROSE',
                   r'ALDI\b',r'LIDL\b',r'MARKS\s*&\s*SPENCER.*FOOD',r'WHOLE\s*FOODS',
                   r'TRADER\s*JOE',r'KROGER',r'WALMART',r'COSTCO',r'BIGBASKET',
                   r'DMART\b',r'RELIANCE\s*FRESH',r'MORE\s*SUPERMARKET',
                   r'WOOLWORTHS',r'COLES\b',r'IGA\b',r'SUPERMARKET',r'GROCERY',
                   r'FARMER.*MARKET',r'KIRANA',r'INSTAMART',r'ZEPTO',r'GROFERS']),
    ('Shopping', [r'AMAZON\b',r'EBAY\b',r'FLIPKART',r'MYNTRA',r'ASOS\b',r'ZARA\b',
                  r'H&M\b',r'H AND M',r'HENNES',r'PRIMARK',r'UNIQLO',r'NEXT\b',
                  r'MARKS\s*&\s*SPENCER',r'JOHN\s*LEWIS',r'BOOTS\b',r'SUPERDRUG',
                  r'DEBENHAMS',r'GAP\b',r'NIKE\b',r'ADIDAS',r'PUMA\b',r'DECATHLON',
                  r'SPORTS\s*DIRECT',r'NYKAA',r'MEESHO',r'AJIO\b',r'SHEIN\b',
                  r'PRETTYLITTLETHING',r'BOOHOO',r'REVOLVE\b',r'SHOPIFY',
                  r'EMPORIUM',r'TATACLIQ',r'SNAPDEAL']),
    ('Travel & Transport', [r'UBER\b',r'LYFT\b',r'OLA\b',r'RAPIDO',r'BOLT\b',
                             r'CABIFY',r'GOJEK',r'GRAB\b',r'DIDI\b',
                             r'RYANAIR',r'EASYJET',r'BRITISH\s*AIRWAYS',r'EMIRATES',
                             r'QATAR\s*AIRWAYS',r'LUFTHANSA',r'INDIGO\b',r'SPICEJET',
                             r'AIR\s*INDIA',r'VISTARA',r'UNITED\s*AIRLINES',r'DELTA\b',
                             r'AMERICAN\s*AIRLINES',r'SOUTHWEST',r'QANTAS',
                             r'BOOKING\.COM',r'AIRBNB',r'HOTELS\.COM',r'EXPEDIA',
                             r'MAKEMYTRIP',r'GOIBIBO',r'CLEARTRIP',r'IXIGO',
                             r'OYO\b',r'TREEBO',r'IRCTC',r'TRAINLINE',r'NATIONAL\s*RAIL',
                             r'TFL\b',r'TRANSPORT\s*FOR\s*LONDON',r'METRO\b',
                             r'OYSTER\b',r'BUS\s*PASS',r'REDBUS',r'TAXI\b',r'CAB\b',
                             r'FLIGHT\b',r'AIRLINE',r'AIRPORT',r'PARKING',
                             r'BMTC\b',r'KSRTC',r'GSRTC',r'TSRTC',r'MSRTC',r'NMMT',
                             r'EMT\s*FLIGHT',r'FLIGHT\s*VIA',r'SB\s*GI\s*FLIGHT',
                             r'HOTEL\b',r'RESORT\b',r'SEA\s*HILLS',r'WATER\s*SPORT',
                             r'REGENTA',r'PREMKUMAR\s*TRAVEL',r'TRAVELS\b']),
    ('Subscriptions', [r'NETFLIX',r'SPOTIFY',r'APPLE\s*(MUSIC|TV|ONE|MEDIA)',
                       r'UPI.*APPLE\b',r'APPLE\s*MEDIA',
                       r'AMAZON\s*PRIME',r'DISNEY\+',r'HBO\b',r'HULU\b',r'PARAMOUNT',
                       r'PEACOCK',r'YOUTUBE\s*PREMIUM',r'GOOGLE\s*ONE',r'GOOGLE\s*INDIA',
                       r'ICLOUD',r'MICROSOFT\s*365',r'OFFICE\s*365',r'ADOBE\b',r'CANVA\b',
                       r'NOTION\b',r'DROPBOX',r'ZOOM\b',r'SLACK\b',r'CLAUDE',
                       r'ANTHROPIC',r'OPENAI',r'CHATGPT',r'GITHUB\b',r'HOTSTAR',
                       r'SONY\s*LIV',r'ZEE5',r'SUBSCRIPTION',r'MEMBERSHIP']),
    ('Telecom', [r'VODAFONE',r'O2\b',r'EE\b',r'THREE\b',r'SKY\b',r'BT\b',r'VIRGIN\s*MEDIA',
                 r'T-MOBILE',r'VERIZON',r'AT&T',r'COMCAST',r'AIRTEL',r'BSNL',r'JIO\b',
                 r'MOBILE\s*BILL',r'PHONE\s*BILL',r'BROADBAND',r'INTERNET\s*BILL',
                 r'DATA\s*PLAN',r'POSTPAID',r'PREPAID',r'RECHARGE']),
    ('Health & Fitness', [r'PHARMACY',r'CHEMIST',r'BOOTS\s*PHARMACY',r'CVS\b',r'WALGREEN',
                           r'NHS\b',r'DOCTOR',r'CLINIC',r'HOSPITAL',r'DENTAL',r'OPTICIAN',
                           r'PHYSIO',r'THERAPY',r'APOLLO\b',r'MEDPLUS',r'NETMEDS',r'1MG\b',
                           r'CULT\.FIT',r'CUREFIT',r'GYM\b',r'FITNESS',r'YOGA\b',
                           r'PILATES',r'NUFFIELD',r'PURE\s*GYM',r'PLANET\s*FITNESS',
                           r'HEALTH\s*FOOD',r'SUPPLEMENT',r'VITAMIN',
                           r'HORIZON\s*MEDICAL',r'MEDICAL\s*CENT',r'FITSHIT',
                           r'PHARMEASY',r'PRACTO',r'LYBRATE',r'CALISTHEN']),
    ('Entertainment', [r'BOOKMYSHOW',r'TICKETMASTER',r'AXS\b',r'EVENTBRITE',r'DICE\b',
                       r'VUE\b',r'ODEON',r'CINEWORLD',r'PVR\b',r'INOX\b',r'CINEPOLIS',
                       r'CINEMA',r'THEATRE',r'MUSEUM',r'GALLERY',r'CONCERT',r'GIG\b',
                       r'BOWLING',r'ESCAPE\s*ROOM',r'GAMING',r'PLAYSTATION',r'XBOX',
                       r'NINTENDO',r'STEAM\b',r'APP\s*STORE',r'PLAY\s*STORE']),
    ('Electronics', [r'APPLE\s*STORE',r'CURRYS\b',r'ARGOS\b',r'PC\s*WORLD',
                     r'CROMA',r'RELIANCE\s*DIGITAL',r'VIJAY\s*SALES',
                     r'BEST\s*BUY',r'NEWEGG',r'LAPTOP',r'SMARTPHONE',r'TABLET',
                     r'HEADPHONE',r'EARPHONE',r'CHARGER',r'ELECTRONIC']),
    ('Fuel', [r'PETROL\b',r'DIESEL\b',r'FUEL\b',r'GAS\s*STATION',r'SHELL\b',
              r'BP\b',r'ESSO\b',r'TEXACO',r'MOBIL\b',r'CHEVRON',
              r'IOCL',r'BPCL',r'HPCL',r'EXXON']),
    ('Education', [r'COURSERA',r'UDEMY',r'SKILLSHARE',r'LINKEDIN\s*LEARNING',
                   r'PLURALSIGHT',r'UDACITY',r'CODECADEMY',r'KHAN\s*ACADEMY',
                   r'UNACADEMY',r'BYJU',r'TUITION',r'SCHOOL\s*FEE',r'COLLEGE\s*FEE',
                   r'UNIVERSITY',r'COURSE\s*FEE',r'EXAM\s*FEE']),
    ('Insurance', [r'INSURANCE',r'INSURE',r'ASSURANCE',r'PREMIUM\s*PAID',
                   r'POLICY\s*RENEWAL',r'AVIVA\b',r'AXA\b',r'ZURICH\b',r'LIC\b',
                   r'PRUDENTIAL',r'ALLIANZ',r'LEGAL\s*&\s*GENERAL']),
    ('Luxury', [r'LOUIS\s*VUITTON',r'GUCCI',r'PRADA',r'HERMES',r'CHANEL',
                r'ROLEX',r'OMEGA\b',r'CARTIER',r'TIFFANY',r'BULGARI',
                r'HARVEY\s*NICHOLS',r'HARRODS',r'SELFRIDGES',r'SAKS',r'NEIMAN',
                r'TANISHQ',r'ETHOS',r'JEWEL',r'DIAMOND',r'LUXURY',r'WATCH\b']),
    ('Personal Care', [r'SALON\b',r'BARBER',r'HAIRDRESS',r'SPA\b',r'MASSAGE',
                       r'MANICURE',r'PEDICURE',r'BEAUTY',r'GROOMING',
                       r'SUPERDRUG',r'LUSH\b',r'THE\s*BODY\s*SHOP',
                       r'SEPHORA',r'ULTA\b',r'NYKAA.*BEAUTY',r'LAKME',
                       r'HEALTH\s*AND\s*GLOW',r'HEALTH\s*&\s*GLOW',
                       r'DRSHETHS',r'DR\s*SHETH',r'FRESH\s*SIGNATURE',
                       r'MAMAEARTH',r'MINIMALIST',r'MCAFFEINE',r'PLUM\b',
                       r'BEARDO',r'MAN\s*COMPANY',r'BOMBAY\s*SHAVING']),
    ('Utilities', [r'ELECTRICITY',r'GAS\s*BILL',r'WATER\s*BILL',r'COUNCIL\s*TAX',
                   r'UTILITY',r'BRITISH\s*GAS',r'EON\b',r'EDF\b',r'OCTOPUS\s*ENERGY',
                   r'BESCOM',r'MSEDCL',r'BSES',r'CESC\b',r'PROPERTY\s*TAX',
                   r'LIVPURE',r'KENT\b',r'AQUAGUARD']),
]

def categorize(desc):
    d = desc.upper()
    for category, patterns in CATEGORY_RULES:
        for p in patterns:
            if re.search(p, d):
                return category
    # EMI fallback — strip prefix and re-match
    if re.search(r'\bEMI\b|\bINSTALMENT\b|\bINSTALLMENT\b', d):
        stripped = re.sub(r'\bEMI\b|\bINSTALMENT\b|\bINSTALLMENT\b', '', d).strip()
        for category, patterns in CATEGORY_RULES:
            for p in patterns:
                if re.search(p, stripped):
                    return category
        return 'Shopping'
    return 'Miscellaneous'

# ── Excel helpers ─────────────────────────────────────────────────────────────

def fmt_currency(v, sym): return f"{sym}{v:,.0f}"

def sec_header(ws, row, col_range, text, color):
    start, end = col_range
    ws.merge_cells(f'{start}{row}:{end}{row}')
    ws[f'{start}{row}'] = text
    ws[f'{start}{row}'].font = Font(name='Arial',size=12,bold=True,color=WHITE)
    ws[f'{start}{row}'].fill = fill(color)
    ws[f'{start}{row}'].alignment = lft()

def build_dashboard(wb, sp, months_order, user_name, sym):
    ws = wb.active; ws.title = "📊 Dashboard"
    ws.sheet_view.showGridLines = False
    for col,w in zip('ABCDEFGH',[2,28,18,18,18,18,18,3]):
        ws.column_dimensions[col].width = w

    ws.merge_cells('B1:G3')
    period = f"{months_order[0]} – {months_order[-1]}" if len(months_order)>1 else (months_order[0] if months_order else '')
    ws['B1'] = f'💰  {user_name.upper()} — SPENDING ANALYSIS  |  {period}'
    ws['B1'].font = Font(name='Arial',size=16,bold=True,color=WHITE)
    ws['B1'].fill = fill(NAVY); ws['B1'].alignment = ctr()

    total = sp['Amount'].sum(); avg_m = total/max(len(months_order),1)
    mt = sp.groupby('Month')['Amount'].sum()
    peak = mt.idxmax() if not mt.empty else '-'; peak_amt = mt.max() if not mt.empty else 0
    top_cat = sp.groupby('Category')['Amount'].sum().idxmax() if not sp.empty else '-'
    wknd_share = sp[sp['IsWeekend']]['Amount'].sum()/total*100 if total>0 else 0

    kpis = [
        ("💳 Total Spend", fmt_currency(total,sym), TEAL),
        ("📅 Avg / Month", fmt_currency(avg_m,sym), NAVY),
        ("🏆 Peak Month", f"{peak}\n{fmt_currency(peak_amt,sym)}", ORANGE),
        ("🔝 Top Category", top_cat, PURPLE),
        ("🧾 Transactions", str(len(sp)), GREEN),
        ("🌙 Weekend %", f"{wknd_share:.1f}%", RED),
    ]
    r = 5
    for col,(title,val,color) in zip('BCDEFG',kpis):
        ws.merge_cells(f'{col}{r}:{col}{r+1}'); ws.merge_cells(f'{col}{r+2}:{col}{r+3}')
        c1=ws[f'{col}{r}']; c1.value=title; c1.font=Font(name='Arial',size=10,bold=True,color=WHITE); c1.fill=fill(color); c1.alignment=ctr()
        c2=ws[f'{col}{r+2}']; c2.value=val; c2.font=Font(name='Arial',size=12,bold=True,color=color); c2.fill=fill(LGRAY); c2.alignment=ctr()
        for i in range(r,r+4): ws[f'{col}{i}'].border=bdr()

    r=10; show_months=months_order[:4]
    sec_header(ws, r, ('B','G'), '📅  MONTHLY SPENDING BY CATEGORY', TEAL); r+=1
    for col,h in zip('BCDEFG',['Category']+show_months+(['TOTAL'] if len(show_months)<5 else [])):
        c=ws[f'{col}{r}']; c.value=h; c.font=hf(10); c.fill=fill(NAVY); c.alignment=ctr(); c.border=bdr()
    r+=1
    mpivot=sp.pivot_table(index='Category',columns='Month',values='Amount',aggfunc='sum',fill_value=0)
    mpivot=mpivot.reindex(columns=[m for m in show_months if m in mpivot.columns],fill_value=0)
    mpivot['TOTAL']=mpivot.sum(axis=1); mpivot=mpivot.sort_values('TOTAL',ascending=False)
    alt=False
    for cat,row_data in mpivot.iterrows():
        bg=LGRAY if alt else WHITE
        ws[f'B{r}']=cat; ws[f'B{r}'].font=bf(10,bold=True); ws[f'B{r}'].fill=fill(bg); ws[f'B{r}'].alignment=lft(); ws[f'B{r}'].border=bdr()
        for col,m in zip('CDEF',show_months):
            v=row_data.get(m,0); ws[f'{col}{r}']=v if v else '-'
            ws[f'{col}{r}'].font=bf(10); ws[f'{col}{r}'].fill=fill(bg); ws[f'{col}{r}'].alignment=rgt(); ws[f'{col}{r}'].border=bdr()
            if v>0: ws[f'{col}{r}'].number_format=f'"{sym}"#,##0'
        ws[f'G{r}']=row_data['TOTAL']; ws[f'G{r}'].font=bf(10,bold=True); ws[f'G{r}'].fill=fill(bg)
        ws[f'G{r}'].alignment=rgt(); ws[f'G{r}'].border=bdr(); ws[f'G{r}'].number_format=f'"{sym}"#,##0'
        alt=not alt; r+=1
    for col in 'BCDEFG':
        ws[f'{col}{r}'].font=hf(10,color='FF000000'); ws[f'{col}{r}'].fill=fill(GOLD); ws[f'{col}{r}'].border=bdr()
    ws[f'B{r}']='TOTAL'; ws[f'B{r}'].alignment=lft()
    for col,m in zip('CDEF',show_months):
        ws[f'{col}{r}']=mpivot[m].sum() if m in mpivot.columns else 0
        ws[f'{col}{r}'].number_format=f'"{sym}"#,##0'; ws[f'{col}{r}'].alignment=rgt()
    ws[f'G{r}']=mpivot['TOTAL'].sum(); ws[f'G{r}'].number_format=f'"{sym}"#,##0'; ws[f'G{r}'].alignment=rgt()

def build_transactions(wb, df, sym):
    ws=wb.create_sheet("📁 All Transactions"); ws.sheet_view.showGridLines=False; ws.freeze_panes='A2'
    headers=['Date','Description',f'Amount ({sym})','Category','Source','Month','Day','Weekend?']
    widths=[14,55,16,22,30,14,14,10]
    for i,(h,w) in enumerate(zip(headers,widths),1):
        ws.column_dimensions[get_column_letter(i)].width=w
        c=ws.cell(1,i,h); c.font=hf(10); c.fill=fill(NAVY); c.alignment=ctr(); c.border=bdr()
    for idx,row in df.sort_values('Date').iterrows():
        r=idx+2; bg=CAT_COLORS.get(row['Category'],'FFFFFFFF')
        vals=[row['Date'].strftime('%d-%b-%Y'),row['Description'],row['Amount'],
              row['Category'],row['Source'],row['Month'],row['Date'].strftime('%A'),'Yes' if row['IsWeekend'] else 'No']
        for i,v in enumerate(vals,1):
            c=ws.cell(r,i,v); c.font=bf(9); c.fill=fill(bg); c.border=bdr()
            c.alignment=rgt() if i==3 else lft()
        ws.cell(r,3).number_format=f'"{sym}"#,##0.00'
    ws.auto_filter.ref=f"A1:{get_column_letter(len(headers))}1"

def build_category(wb, sp, sym):
    ws=wb.create_sheet("🗂️ Category Breakdown"); ws.sheet_view.showGridLines=False
    for col,w in zip('ABCDEFGH',[2,30,18,12,16,18,18,3]):
        ws.column_dimensions[col].width=w
    ws.merge_cells('B1:G2'); ws['B1']='🗂️  CATEGORY-WISE SPENDING BREAKDOWN'
    ws['B1'].font=Font(name='Arial',size=14,bold=True,color=WHITE); ws['B1'].fill=fill(TEAL); ws['B1'].alignment=ctr()
    r=4
    for col,h in zip('BCDEFG',['Category',f'Total Spend ({sym})','% of Total','Transactions',f'Avg Ticket ({sym})','Biggest Month']):
        c=ws[f'{col}{r}']; c.value=h; c.font=hf(10); c.fill=fill(NAVY); c.alignment=ctr(); c.border=bdr()
    r+=1
    cs=sp.groupby('Category').agg(Total=('Amount','sum'),Count=('Amount','count')).sort_values('Total',ascending=False)
    cs['Pct']=cs['Total']/cs['Total'].sum(); cs['Avg']=cs['Total']/cs['Count']
    cr_start=r+len(cs)+2
    for cat,crow in cs.iterrows():
        bm=sp[sp['Category']==cat].groupby('Month')['Amount'].sum()
        bg=CAT_COLORS.get(cat,'FFFFFFFF')
        for col,val in zip('BCDEFG',[cat,crow['Total'],crow['Pct'],int(crow['Count']),crow['Avg'],bm.idxmax() if not bm.empty else '-']):
            c=ws[f'{col}{r}']; c.value=val; c.font=bf(10,bold=(col=='B')); c.fill=fill(bg)
            c.alignment=rgt() if col in 'CDEF' else lft(); c.border=bdr()
        ws[f'C{r}'].number_format=f'"{sym}"#,##0'; ws[f'D{r}'].number_format='0.0%'; ws[f'F{r}'].number_format=f'"{sym}"#,##0'; r+=1
    for col in 'BCDEFG':
        ws[f'{col}{r}'].font=hf(10,color='FF000000'); ws[f'{col}{r}'].fill=fill(GOLD); ws[f'{col}{r}'].border=bdr()
    ws[f'B{r}']='TOTAL'; ws[f'B{r}'].alignment=lft()
    ws[f'C{r}']=cs['Total'].sum(); ws[f'C{r}'].number_format=f'"{sym}"#,##0'; ws[f'C{r}'].alignment=rgt()
    ws[f'D{r}']=1.0; ws[f'D{r}'].number_format='0.0%'; ws[f'D{r}'].alignment=rgt()
    ws[f'E{r}']=int(cs['Count'].sum()); ws[f'E{r}'].alignment=rgt()
    r+=2; cr_s=r; ws[f'B{r}']='Category'; ws[f'C{r}']='Amount'; r+=1
    for cat,crow in cs.iterrows(): ws[f'B{r}']=cat; ws[f'C{r}']=crow['Total']; r+=1
    pie=PieChart(); pie.title="Spending by Category"; pie.style=10
    pie.add_data(Reference(ws,min_col=3,min_row=cr_s,max_row=r-1),titles_from_data=True)
    pie.set_categories(Reference(ws,min_col=2,min_row=cr_s+1,max_row=r-1))
    pie.width=18; pie.height=14; ws.add_chart(pie,f'B{cr_s}')

def build_trends(wb, sp, months_order, sym):
    ws=wb.create_sheet("📅 Monthly Trends"); ws.sheet_view.showGridLines=False
    for col,w in zip('ABCDEFGH',[2,25,18,18,18,18,18,3]): ws.column_dimensions[col].width=w
    ws.merge_cells('B1:G2'); ws['B1']='📅  MONTHLY SPENDING TRENDS'
    ws['B1'].font=Font(name='Arial',size=14,bold=True,color=WHITE); ws['B1'].fill=fill(ORANGE); ws['B1'].alignment=ctr()
    r=4; show_months=months_order[:4]
    mt=sp.groupby('Month')['Amount'].sum().reindex(show_months,fill_value=0)
    mc=sp.groupby('Month')['Amount'].count().reindex(show_months,fill_value=0)
    ma=sp.groupby('Month')['Amount'].mean().reindex(show_months,fill_value=0)
    for col,h in zip('BCDEFG',['Metric']+show_months+['']):
        c=ws[f'{col}{r}']; c.value=h; c.font=hf(10); c.fill=fill(NAVY); c.alignment=ctr(); c.border=bdr()
    r+=1
    for label,series,fmt,bg in [('Total Spend',mt,f'"{sym}"#,##0',LGRAY),
                                  ('Transactions',mc,'#,##0',WHITE),
                                  ('Avg Transaction',ma,f'"{sym}"#,##0',LGRAY)]:
        ws[f'B{r}']=label; ws[f'B{r}'].font=bf(10,bold=True); ws[f'B{r}'].fill=fill(bg); ws[f'B{r}'].alignment=lft(); ws[f'B{r}'].border=bdr()
        for col,m in zip('CDEF',show_months):
            c=ws[f'{col}{r}']; c.value=series.get(m,0); c.font=bf(10); c.fill=fill(bg); c.alignment=rgt(); c.border=bdr(); c.number_format=fmt
        r+=1
    r+=1; sec_header(ws,r,('B','D'),'🌙 Weekday vs Weekend Spending',PURPLE); r+=1
    for col,h in zip('BCD',['Type','Total Spend','% of Total']):
        c=ws[f'{col}{r}']; c.value=h; c.font=hf(10); c.fill=fill(NAVY); c.alignment=ctr(); c.border=bdr()
    r+=1
    wday=sp.groupby('IsWeekend')['Amount'].sum(); total_wk=wday.sum()
    for is_wknd,label in [(False,'Weekday (Mon–Fri)'),(True,'Weekend (Sat–Sun)')]:
        v=wday.get(is_wknd,0)
        for col,val in zip('BCD',[label,v,v/total_wk if total_wk>0 else 0]):
            c=ws[f'{col}{r}']; c.value=val; c.font=bf(10); c.fill=fill(LGRAY)
            c.alignment=rgt() if col in 'CD' else lft(); c.border=bdr()
        ws[f'C{r}'].number_format=f'"{sym}"#,##0'; ws[f'D{r}'].number_format='0.0%'; r+=1
    r+=2; cr=r; ws[f'B{r}']='Month'; ws[f'C{r}']='Total'; r+=1
    for m in show_months: ws[f'B{r}']=m; ws[f'C{r}']=mt.get(m,0); r+=1
    bar=BarChart(); bar.title="Monthly Total Spend"; bar.y_axis.title=f"Amount ({sym})"; bar.style=10; bar.type="col"
    bar.add_data(Reference(ws,min_col=3,min_row=cr,max_row=cr+len(show_months)),titles_from_data=True)
    bar.set_categories(Reference(ws,min_col=2,min_row=cr+1,max_row=cr+len(show_months)))
    bar.width=18; bar.height=12; ws.add_chart(bar,f'E{cr}')

def build_merchants(wb, sp, sym):
    ws=wb.create_sheet("🏪 Top Merchants"); ws.sheet_view.showGridLines=False
    for col,w in zip('ABCDEFG',[2,50,18,12,18,22,3]): ws.column_dimensions[col].width=w
    ws.merge_cells('B1:F2'); ws['B1']='🏪  TOP MERCHANTS BY SPEND'
    ws['B1'].font=Font(name='Arial',size=14,bold=True,color=WHITE); ws['B1'].fill=fill(ORANGE); ws['B1'].alignment=ctr()
    r=4
    for col,h in zip('BCDEF',['Merchant',f'Total Spend ({sym})','Times',f'Avg Ticket ({sym})','Category']):
        c=ws[f'{col}{r}']; c.value=h; c.font=hf(10); c.fill=fill(NAVY); c.alignment=ctr(); c.border=bdr()
    r+=1
    top_m=sp.groupby('Description').agg(Total=('Amount','sum'),Count=('Amount','count')).sort_values('Total',ascending=False).head(20)
    top_m['Avg']=top_m['Total']/top_m['Count']
    top_m['Cat']=[sp[sp['Description']==d]['Category'].iloc[0] for d in top_m.index]
    medal={0:'🥇',1:'🥈',2:'🥉'}
    for i,(desc,mrow) in enumerate(top_m.iterrows()):
        bg=['FFFFF9C4','FFF3E5F5','FFE8F5E9'][i] if i<3 else (LGRAY if i%2==0 else WHITE)
        for col,val in zip('BCDEF',[f"{medal.get(i,f'#{i+1}  ')}  {desc}",mrow['Total'],int(mrow['Count']),mrow['Avg'],mrow['Cat']]):
            c=ws[f'{col}{r}']; c.value=val; c.font=bf(10,bold=(i<3)); c.fill=fill(bg)
            c.alignment=rgt() if col in 'CDE' else lft(); c.border=bdr()
        ws[f'C{r}'].number_format=f'"{sym}"#,##0'; ws[f'E{r}'].number_format=f'"{sym}"#,##0'; r+=1
    r+=2; sec_header(ws,r,('B','F'),'🔄  RECURRING PAYMENTS (2+ times)',TEAL); r+=1
    for col,h in zip('BCDE',['Merchant','Frequency','Total Paid','Category']):
        c=ws[f'{col}{r}']; c.value=h; c.font=hf(10); c.fill=fill(NAVY); c.alignment=ctr(); c.border=bdr()
    r+=1
    recur=sp.groupby('Description').filter(lambda x:len(x)>=2).groupby('Description').agg(
        Count=('Amount','count'),Total=('Amount','sum')).sort_values('Count',ascending=False).head(12)
    if not recur.empty:
        recur['Cat']=[sp[sp['Description']==d]['Category'].iloc[0] for d in recur.index]
        for desc,rrow in recur.iterrows():
            for col,val in zip('BCDE',[desc,f"{int(rrow['Count'])}x",rrow['Total'],rrow['Cat']]):
                c=ws[f'{col}{r}']; c.value=val; c.font=bf(10); c.fill=fill(LGRAY)
                c.alignment=rgt() if col=='C' else lft(); c.border=bdr()
            ws[f'C{r}'].number_format=f'"{sym}"#,##0'; r+=1

def build_insights(wb, sp, sym):
    ws=wb.create_sheet("💡 Insights & Budget"); ws.sheet_view.showGridLines=False
    for col,w in zip('ABCDEFG',[2,45,18,18,18,18,3]): ws.column_dimensions[col].width=w
    ws.merge_cells('B1:F2'); ws['B1']='💡  BEHAVIOUR INSIGHTS & BUDGET PLAN'
    ws['B1'].font=Font(name='Arial',size=14,bold=True,color=WHITE); ws['B1'].fill=fill(NAVY); ws['B1'].alignment=ctr()

    total=sp['Amount'].sum(); n_months=max(sp['Month'].nunique(),1)
    top3=sp.groupby('Category')['Amount'].sum().sort_values(ascending=False).head(3)
    misc=sp[sp['Category']=='Miscellaneous'].sort_values('Amount',ascending=False).head(5)

    overspend=[f"{cat}: {fmt_currency(amt,sym)} ({amt/total*100:.1f}% of spend) — review if aligned with priorities"
               for cat,amt in top3.items()]
    misc_items=[f"  • {r['Description'][:45]} — {fmt_currency(r['Amount'],sym)}" for _,r in misc.iterrows()]

    sections=[
        ("🚨 BIGGEST SPEND AREAS", RED, overspend),
        ("🔎 UNCATEGORISED — REVIEW THESE", ORANGE,
         ["These merchants weren't recognised — tell your analyst what they are:"] + misc_items if misc_items else ["None — great categorisation!"]),
        ("💰 SAVINGS OPPORTUNITIES", TEAL, [
            "Audit all subscriptions — cancel anything unused in the last 30 days",
            "Batch grocery shopping weekly instead of daily convenience orders (saves 15-20%)",
            "Pre-book travel 3-4 weeks ahead — last-minute bookings cost significantly more",
            "Set a monthly shopping cap and use a 48-hour wish-list rule before buying",
            "Check if any EMI/instalment plans overlap — total should be <30% of income",
        ]),
        ("✅ HABITS TO KEEP", GREEN, [
            "Tracking every month is the single most powerful financial habit — keep going",
            "Protect your investment contributions — don't let lifestyle creep reduce them",
            "If weekend spend is under 30% of total, your impulse control is healthy",
        ]),
    ]
    r=4
    for title,color,points in sections:
        ws.merge_cells(f'B{r}:F{r}'); ws[f'B{r}']=title
        ws[f'B{r}'].font=Font(name='Arial',size=11,bold=True,color=WHITE); ws[f'B{r}'].fill=fill(color); ws[f'B{r}'].alignment=lft(); r+=1
        for pt in points:
            ws.merge_cells(f'B{r}:F{r}'); ws[f'B{r}']=f"   •  {pt.strip()}"
            ws[f'B{r}'].font=bf(10); ws[f'B{r}'].fill=fill(LGRAY if r%2==0 else WHITE)
            ws[f'B{r}'].alignment=Alignment(horizontal='left',vertical='center',wrap_text=True)
            ws[f'B{r}'].border=bdr(); ws.row_dimensions[r].height=28; r+=1
        r+=1
    r+=1; ws.merge_cells(f'B{r}:F{r}'); ws[f'B{r}']=f'📋  SUGGESTED MONTHLY BUDGET  ({sym})'
    ws[f'B{r}'].font=Font(name='Arial',size=12,bold=True,color=WHITE); ws[f'B{r}'].fill=fill(NAVY); ws[f'B{r}'].alignment=lft(); r+=1
    for col,h in zip('BCDEF',['Category','Avg Actual/Month','Suggested Budget','Potential Saving','Tip']):
        c=ws[f'{col}{r}']; c.value=h; c.font=hf(10); c.fill=fill(NAVY); c.alignment=ctr(); c.border=bdr()
    r+=1
    cat_avgs=sp.groupby('Category')['Amount'].sum()/n_months; ta=tb=0
    for cat in sorted(cat_avgs.index):
        actual=cat_avgs[cat]; pct=actual/total*100
        budget=actual*0.8 if pct>15 else actual*0.9 if pct>8 else actual
        saving=actual-budget; tip='Reduce by 20%' if pct>15 else 'Reduce by 10%' if pct>8 else 'Maintain'
        bg='FFE8F5E9' if saving>0 else 'FFFFEBEE' if saving<-50 else LGRAY
        for col,val in zip('BCDEF',[cat,round(actual),round(budget),round(saving),tip]):
            c=ws[f'{col}{r}']; c.value=val; c.font=bf(10); c.fill=fill(bg)
            c.alignment=rgt() if col in 'CDE' else lft(); c.border=bdr()
        for col in 'CDE': ws[f'{col}{r}'].number_format=f'"{sym}"#,##0'
        ta+=actual; tb+=budget; r+=1
    for col,val in zip('BCDE',['TOTAL',round(ta),round(tb),round(ta-tb)]):
        c=ws[f'{col}{r}']; c.value=val; c.font=hf(10,color='FF000000'); c.fill=fill(GOLD)
        c.alignment=rgt() if col!='B' else lft(); c.border=bdr()
        if col in 'CDE': ws[f'{col}{r}'].number_format=f'"{sym}"#,##0'
    ws[f'F{r}']=f"Potential saving: {fmt_currency(ta-tb,sym)}/month"
    ws[f'F{r}'].font=Font(name='Arial',size=10,bold=True,color=GREEN); ws[f'F{r}'].fill=fill(GOLD); ws[f'F{r}'].border=bdr()

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser=argparse.ArgumentParser()
    parser.add_argument('--input', required=True)
    parser.add_argument('--output', required=True)
    parser.add_argument('--name', default='Your')
    parser.add_argument('--currency', default=None, help='Currency symbol override (e.g. $ £ € ₹)')
    args=parser.parse_args()

    df=pd.read_csv(args.input, parse_dates=['Date'])

    # Detect currency from CSV if not overridden
    sym = args.currency
    if not sym and 'Currency' in df.columns:
        sym = df['Currency'].iloc[0] if not df['Currency'].empty else '$'
    if not sym: sym = '$'

    df['Category']=df['Description'].apply(categorize)
    df['Month']=df['Date'].dt.strftime('%b %Y')
    df['IsWeekend']=df['Date'].dt.dayofweek>=5
    months_order=sorted(df['Month'].unique(), key=lambda m: pd.to_datetime(m,format='%b %Y'))
    sp=df[~df['Category'].isin(NON_SPEND)].copy().reset_index(drop=True)

    wb=Workbook()
    build_dashboard(wb, sp, months_order, args.name, sym)
    build_transactions(wb, df, sym)
    build_category(wb, sp, sym)
    build_trends(wb, sp, months_order, sym)
    build_merchants(wb, sp, sym)
    build_insights(wb, sp, sym)
    wb.save(args.output)

    print(f"✅ Report saved: {args.output}")
    print(f"   📊 {len(df)} transactions | {fmt_currency(sp['Amount'].sum(),sym)} spend | {len(months_order)} month(s) | currency: {sym}")
    misc=len(sp[sp['Category']=='Miscellaneous'])
    if misc: print(f"   ⚠️  {misc} transactions in Miscellaneous — review in Insights sheet")

if __name__=='__main__':
    main()
