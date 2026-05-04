#!/usr/bin/env python3
"""
Generate spending analysis Excel report from transaction CSV.
Fully generic — works for any user, any bank, any set of transactions.
"""
import argparse, re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference

# ── Styles ──────────────────────────────────────────────────────────────────
NAVY   = "FF1F3864"; TEAL   = "FF0097A7"; ORANGE = "FFFF6F00"; GOLD = "FFFFC107"
WHITE  = "FFFFFFFF"; LGRAY  = "FFF5F5F5"; RED    = "FFD32F2F"
GREEN  = "FF2E7D32"; PURPLE = "FF6A1B9A"; DBLUE  = "FF1565C0"

def fill(h): return PatternFill("solid", start_color=h, end_color=h)
def hf(size=11, bold=True, color=WHITE): return Font(name='Arial', size=size, bold=bold, color=color)
def bf(size=10, bold=False, color='FF000000'): return Font(name='Arial', size=size, bold=bold, color=color)
def ctr(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
def lft(): return Alignment(horizontal='left', vertical='center', wrap_text=True)
def rgt(): return Alignment(horizontal='right', vertical='center')
def bdr():
    s = Side(style='thin', color='FFCCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)
def inr(v): return f"₹{v:,.0f}"

CAT_COLORS = {
    'Food & Dining':     'FFE3F2FD', 'Groceries':         'FFE8F5E9',
    'Shopping':          'FFFFF3E0', 'Travel & Transport': 'FFF3E5F5',
    'Subscriptions':     'FFFCE4EC', 'Health & Fitness':  'FFE0F2F1',
    'Entertainment':     'FFFFF8E1', 'Electronics':       'FFE8EAF6',
    'Luxury & Jewelry':  'FFF8F0FF', 'Rent & Housing':    'FFEFEBE9',
    'Utilities':         'FFE0F7FA', 'Personal Care':     'FFFFF9C4',
    'Telecom':           'FFF1F8E9', 'Fuel':              'FFFFF3E0',
    'Education':         'FFE8EAF6', 'Insurance':         'FFE0F7FA',
    'Investments':       'FFE8F5E9', 'Transfers':         'FFFAFAFA',
    'Miscellaneous':     'FFFAFAFA',
}
NON_SPEND = {'Investments', 'Transfers'}

# ── Categorisation ──────────────────────────────────────────────────────────
CATEGORY_RULES = [
    # Investments
    ('Investments', ['ZERODHA','GROWW','KUVERA','COIN BY ZERODHA','SMALLCASE','NACH.*MF',
                     'NACH.*MUT','MUTUAL FUND','SIP','NIFTY','ICICI DIRECT','HDFC AMC',
                     'SBI AMC','AXIS AMC','NIPPON','MIRAE','PARAG PARIKH','UTI AMC',
                     'INDmoney','ANGEL BROKING','UPSTOX','5PAISA']),
    # Rent & Housing
    ('Rent & Housing', ['RENT','HOUSE','FLAT','PG ','PAYING GUEST','MAINTENANCE',
                        'SOCIETY','HOUSING','LANDLORD','OWNER','NOBROKER','MAGICBRICKS',
                        '99ACRES','NESTAWAY','STANZA','COLIVE','YOUWE']),
    # Transfers (skip from spending)
    ('Transfers', ['IMPS TO SELF','NEFT TO SELF','UPI.*SELF','TRANSFER TO OWN',
                   'OWN ACCOUNT']),
    # Food & Dining
    ('Food & Dining', ['ZOMATO','SWIGGY','BLINKIT','DUNZO','BUNDL','ZEPTO',
                       'BIGBASKET.*EXPRESS','INSTAMART','DOMINOS','PIZZA HUT','KFC',
                       'MCDONALDS','BURGER KING','SUBWAY','STARBUCKS','CAFE COFFEE DAY',
                       'COSTA COFFEE','CHAAYOS','CHAI POINT','BARISTA','THIRD WAVE',
                       'RESTAURANT','CAFE','DHABA','BIRYANI','CANTEEN','FOOD COURT',
                       'DINING','EATERY','KITCHEN','BAKERY','SWEET SHOP','JUICE',
                       'TEA STALL','HOTEL.*FOOD','TASTE','EAT','MESS ']),
    # Groceries
    ('Groceries', ['BIGBASKET','RELIANCE FRESH','DMART','MORE SUPERMARKET',
                   'NATURE.S BASKET','STAR BAZAAR','SPAR','LULU','HYPER','GROCERY',
                   'SUPER MARKET','SUPERMARKET','KIRANA','PROVISIONS','VEGETABLES',
                   'FRUITS','MILK','DAIRY','MARUTHI STORE','RATNADEEP']),
    # Shopping
    ('Shopping', ['AMAZON','FLIPKART','MYNTRA','MEESHO','NYKAA','AJIO','TATA CLIQ',
                  'SNAPDEAL','SHOPCLUES','LIMEROAD','URBANIC','BEWAKOOF','MANYAVAR',
                  'FABINDIA','W FOR WOMAN','BIBA','PANTALOONS','SHOPPERS STOP',
                  'LIFESTYLE','WESTSIDE','H&M','ZARA','UNIQLO','MARKS AND SPENCER',
                  'GAP ','FOREVER 21','AND ','GLOBAL DESI','RELIANCE TRENDS',
                  'MAX FASHION','V-MART','VISHAL MEGA','BIG BAZAAR','DECATHLON',
                  'NIKE','ADIDAS','PUMA','REEBOK','SKECHERS','BATA','METRO SHOES',
                  'WOODLAND','RED TAPE','LIBERTY SHOES']),
    # Travel & Transport
    ('Travel & Transport', ['UBER','OLA ','RAPIDO','MERU','INDRIVE','PORTER',
                             'IRCTC','TRAIN','RAILWAYS','MAKEMYTRIP','GOIBIBO',
                             'CLEARTRIP','YATRA','IXIGO','EASEMYTRIP','BOOKING.COM',
                             'AIRBNB','OYO','TREEBO','FABHOTEL','ZO ROOMS',
                             'INDIGO','SPICEJET','AIRINDIA','AIR ASIA','VISTARA',
                             'FLIGHT','TICKET','TRAVEL','METRO ','BUS ','AUTO ',
                             'RAPIDO','BMTC','DTC ','BEST BUS','APSRTC','MSRTC',
                             'REDBUS','ABHIBUS','PAYTM.*TRAVEL']),
    # Subscriptions
    ('Subscriptions', ['NETFLIX','AMAZON PRIME','HOTSTAR','DISNEY','SONY LIV',
                       'ZEE5','VOOT','MXPLAYER','JIOCINEMA','APPLE.*MEDIA',
                       'APPLE.*TV','YOUTUBE.*PREMIUM','SPOTIFY','GAANA','WYNK',
                       'JIOSAAVN','GOOGLE ONE','GOOGLE.*STORAGE','MICROSOFT 365',
                       'OFFICE 365','ADOBE','CANVA','NOTION','ZOOM','SLACK',
                       'DROPBOX','ICLOUD','CLAUDE','ANTHROPIC','CHATGPT',
                       'OPENAI','LINKEDIN PREMIUM','SUBSCRIPTION','MEMBERSHIP']),
    # Telecom
    ('Telecom', ['AIRTEL','VODAFONE','VI ','BSNL','JIOMART','RECHARGE',
                 'MOBILE BILL','POSTPAID','PREPAID','DATA PACK','TATA PLAY',
                 'DISH TV','VIDEOCON D2H','SUN DIRECT','HATHWAY','ACT BROADBAND',
                 'SPECTRANET','TIKONA','EXCITEL']),
    # Health & Fitness
    ('Health & Fitness', ['APOLLO','MEDPLUS','NETMEDS','1MG','PHARMEASY',
                           'HEALTHKART','CULT.FIT','CUREFIT','FITTERFLY',
                           'DOCTOR','CLINIC','HOSPITAL','PHARMACY','MEDICAL',
                           'DIAGNOSTIC','WELLNESS','THERAPY','CONSULT',
                           'GYM','FITNESS','YOGA','PILATES','PHYSIOTHERAPY',
                           'PRACTO','TATA 1MG','MFINE','LYBRATE']),
    # Entertainment
    ('Entertainment', ['BOOKMYSHOW','PAYTM INSIDER','DISTRICT','PVR','INOX',
                       'CINEPOLIS','CARNIVAL','MOVIE','CONCERT','EVENT',
                       'AMUSEMENT','THEME PARK','ESCAPE ROOM','BOWLING',
                       'GAMING','PLAY STORE','APP STORE','STEAM']),
    # Electronics
    ('Electronics', ['CROMA','RELIANCE DIGITAL','VIJAY SALES','EZONE','SANGEETHA',
                     'APPLE STORE','SAMSUNG','XIAOMI','REALME','ONEPLUS',
                     'LAPTOP','MOBILE','PHONE','TABLET','EARPHONE','HEADPHONE',
                     'CHARGER','GADGET','ELECTRONIC']),
    # Fuel
    ('Fuel', ['PETROL','DIESEL','FUEL','IOCL','BPCL','HPCL','SHELL','ESSAR OIL',
              'NAYARA','HP PUMP','BP PUMP','SPEED BUNK','RELIANCE BP']),
    # Education
    ('Education', ['COURSERA','UDEMY','UNACADEMY','BYJU','VEDANTU','WHITEHAT',
                   'TOPPR','KHAN ACADEMY','SIMPLILEARN','UPGRAD','GREAT LEARNING',
                   'COLLEGE','SCHOOL','UNIVERSITY','TUITION','CLASSES','COACHING',
                   'EXAM FEE','APPLICATION FEE','ADMISSION']),
    # Insurance
    ('Insurance', ['INSURANCE','IRDAI','LIC ','TATA AIG','BAJAJ ALLIANZ',
                   'STAR HEALTH','NIVA BUPA','CARE HEALTH','MAX BUPA',
                   'HDFC ERGO','ICICI LOMBARD','NEW INDIA','NATIONAL INSURANCE',
                   'TERM PLAN','PREMIUM PAID','POLICY']),
    # Luxury & Jewelry
    ('Luxury & Jewelry', ['TITAN','TANISHQ','KALYAN JEWELLERS','MALABAR',
                           'JOYALUKKAS','ETHOS','HELIOS','FOSSIL','CASIO',
                           'LUXURY','JEWEL','GOLD SHOP','DIAMOND','PLATINUM',
                           'WATCHES','RAYMOND','LOUIS PHILIPPE','VAN HEUSEN',
                           'PARK AVENUE','ALLEN SOLLY','ARROW ']),
    # Personal Care
    ('Personal Care', ['SALON','SPA','PARLOUR','BEAUTY','NAILS','MANICURE',
                       'PEDICURE','HAIRCUT','BARBER','GROOMING','WAXING',
                       'LAKME','VLCC','BODYCRAFT','NATURALS','YLG',
                       'JUST FOR HEARTS','GREEN TRENDS','LOREAL','MAMAEARTH',
                       'MCAFFEINE','PLUM','DOT & KEY','MINIMALIST','PILGRIM']),
    # Utilities
    ('Utilities', ['BESCOM','MSEDCL','TPDDL','BSES','CESC','ELECTRICITY',
                   'WATER BOARD','GAS PIPE','MAHANAGAR GAS','IGL ','ADANI GAS',
                   'BIMAPLAN','MUNICIPAL','BBMP','BMC ','AMC ','PROPERTY TAX',
                   'MAINTENANCE CHARGE']),
]

def categorize(desc):
    d = desc.upper()
    for category, keywords in CATEGORY_RULES:
        for kw in keywords:
            if re.search(kw, d):
                return category
    # EMI detection — try to categorise by merchant inside EMI label
    if 'EMI' in d:
        emi_desc = re.sub(r'EMI\s*', '', d, flags=re.IGNORECASE)
        for category, keywords in CATEGORY_RULES:
            for kw in keywords:
                if re.search(kw, emi_desc):
                    return category
        return 'Shopping'  # Default for unknown EMIs
    return 'Miscellaneous'

# ── Sheet builders ──────────────────────────────────────────────────────────

def build_dashboard(wb, sp, months_order, user_name):
    ws = wb.active; ws.title = "📊 Dashboard"
    ws.sheet_view.showGridLines = False
    for col, w in zip('ABCDEFGH', [2,28,18,18,18,18,18,3]):
        ws.column_dimensions[col].width = w

    ws.merge_cells('B1:G3')
    period = f"{months_order[0]} – {months_order[-1]}" if len(months_order)>1 else months_order[0] if months_order else ''
    ws['B1'] = f'💰  {user_name.upper()} — SPENDING ANALYSIS  |  {period}'
    ws['B1'].font = Font(name='Arial', size=16, bold=True, color=WHITE)
    ws['B1'].fill = fill(NAVY); ws['B1'].alignment = ctr()

    total_sp = sp['Amount'].sum()
    avg_monthly = total_sp / max(len(months_order), 1)
    month_totals = sp.groupby('Month')['Amount'].sum()
    peak = month_totals.idxmax() if not month_totals.empty else '-'
    peak_amt = month_totals.max() if not month_totals.empty else 0
    top_cat = sp.groupby('Category')['Amount'].sum().idxmax() if not sp.empty else '-'
    weekend_share = sp[sp['IsWeekend']]['Amount'].sum() / total_sp * 100 if total_sp > 0 else 0

    kpis = [
        ("💳 Total Spend", inr(total_sp), TEAL),
        ("📅 Avg / Month", inr(avg_monthly), NAVY),
        (f"🏆 Peak Month", f"{peak}\n{inr(peak_amt)}", ORANGE),
        ("🔝 Top Category", top_cat, PURPLE),
        ("🧾 Transactions", str(len(sp)), GREEN),
        ("🌙 Weekend %", f"{weekend_share:.1f}%", RED),
    ]
    r = 5
    for col, (title, val, color) in zip('BCDEFG', kpis):
        ws.merge_cells(f'{col}{r}:{col}{r+1}'); ws.merge_cells(f'{col}{r+2}:{col}{r+3}')
        c1 = ws[f'{col}{r}']; c1.value = title
        c1.font = Font(name='Arial',size=10,bold=True,color=WHITE); c1.fill = fill(color); c1.alignment = ctr()
        c2 = ws[f'{col}{r+2}']; c2.value = val
        c2.font = Font(name='Arial',size=12,bold=True,color=color); c2.fill = fill(LGRAY); c2.alignment = ctr()
        for i in range(r, r+4): ws[f'{col}{i}'].border = bdr()

    r = 10
    ws.merge_cells(f'B{r}:G{r}')
    ws[f'B{r}'] = '📅  MONTHLY SPENDING BY CATEGORY'
    ws[f'B{r}'].font = Font(name='Arial',size=12,bold=True,color=WHITE)
    ws[f'B{r}'].fill = fill(TEAL); ws[f'B{r}'].alignment = lft(); r += 1

    show_months = months_order[:4]
    hdr_cols = ['Category'] + show_months + (['TOTAL'] if len(show_months) < 5 else [])
    for col, h in zip('BCDEFG', hdr_cols):
        c = ws[f'{col}{r}']; c.value = h; c.font = hf(10)
        c.fill = fill(NAVY); c.alignment = ctr(); c.border = bdr()
    r += 1

    mpivot = sp.pivot_table(index='Category', columns='Month', values='Amount', aggfunc='sum', fill_value=0)
    mpivot = mpivot.reindex(columns=[m for m in show_months if m in mpivot.columns], fill_value=0)
    mpivot['TOTAL'] = mpivot.sum(axis=1)
    mpivot = mpivot.sort_values('TOTAL', ascending=False)
    alt = False
    for cat, row_data in mpivot.iterrows():
        bg = LGRAY if alt else WHITE
        ws[f'B{r}'] = cat; ws[f'B{r}'].font = bf(10, bold=True); ws[f'B{r}'].fill = fill(bg)
        ws[f'B{r}'].alignment = lft(); ws[f'B{r}'].border = bdr()
        for col, m in zip('CDEF', show_months):
            v = row_data.get(m, 0)
            ws[f'{col}{r}'] = v if v else '-'; ws[f'{col}{r}'].font = bf(10)
            ws[f'{col}{r}'].fill = fill(bg); ws[f'{col}{r}'].alignment = rgt(); ws[f'{col}{r}'].border = bdr()
            if v > 0: ws[f'{col}{r}'].number_format = '₹#,##0'
        ws[f'G{r}'] = row_data['TOTAL']; ws[f'G{r}'].font = bf(10, bold=True)
        ws[f'G{r}'].fill = fill(bg); ws[f'G{r}'].alignment = rgt()
        ws[f'G{r}'].border = bdr(); ws[f'G{r}'].number_format = '₹#,##0'
        alt = not alt; r += 1
    for col in 'BCDEFG':
        ws[f'{col}{r}'].font = hf(10, color='FF000000'); ws[f'{col}{r}'].fill = fill(GOLD); ws[f'{col}{r}'].border = bdr()
    ws[f'B{r}'] = 'TOTAL'; ws[f'B{r}'].alignment = lft()
    for col, m in zip('CDEF', show_months):
        ws[f'{col}{r}'] = mpivot[m].sum() if m in mpivot.columns else 0
        ws[f'{col}{r}'].number_format = '₹#,##0'
    ws[f'G{r}'] = mpivot['TOTAL'].sum(); ws[f'G{r}'].number_format = '₹#,##0'
    for col in 'CDEFG': ws[f'{col}{r}'].alignment = rgt()

def build_transactions(wb, df):
    ws = wb.create_sheet("📁 All Transactions")
    ws.sheet_view.showGridLines = False; ws.freeze_panes = 'A2'
    headers = ['Date','Description','Amount (₹)','Category','Payment Method','Month','Day','Weekend?']
    widths =  [14, 55, 16, 22, 30, 14, 14, 12]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        c = ws.cell(1, i, h); c.font = hf(10); c.fill = fill(NAVY); c.alignment = ctr(); c.border = bdr()
    for idx, row in df.sort_values('Date').iterrows():
        r = idx + 2; bg = CAT_COLORS.get(row['Category'], 'FFFFFFFF')
        vals = [row['Date'].strftime('%d-%b-%Y'), row['Description'], row['Amount'],
                row['Category'], row['Source'], row['Month'],
                row['Date'].strftime('%A'), 'Yes' if row['IsWeekend'] else 'No']
        for i, v in enumerate(vals, 1):
            c = ws.cell(r, i, v); c.font = bf(9); c.fill = fill(bg)
            c.border = bdr(); c.alignment = rgt() if i == 3 else lft()
        ws.cell(r, 3).number_format = '₹#,##0.00'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

def build_category(wb, sp):
    ws = wb.create_sheet("🗂️ Category Breakdown")
    ws.sheet_view.showGridLines = False
    for col, w in zip('ABCDEFGH', [2,30,18,12,16,18,18,3]):
        ws.column_dimensions[col].width = w
    ws.merge_cells('B1:G2')
    ws['B1'] = '🗂️  CATEGORY-WISE SPENDING BREAKDOWN'
    ws['B1'].font = Font(name='Arial',size=14,bold=True,color=WHITE)
    ws['B1'].fill = fill(TEAL); ws['B1'].alignment = ctr()
    r = 4
    for col, h in zip('BCDEFG', ['Category','Total Spend (₹)','% of Total','Transactions','Avg Ticket (₹)','Biggest Month']):
        c = ws[f'{col}{r}']; c.value = h; c.font = hf(10); c.fill = fill(NAVY); c.alignment = ctr(); c.border = bdr()
    r += 1
    cat_sum = sp.groupby('Category').agg(Total=('Amount','sum'),Count=('Amount','count')).sort_values('Total',ascending=False)
    cat_sum['Pct'] = cat_sum['Total'] / cat_sum['Total'].sum()
    cat_sum['Avg'] = cat_sum['Total'] / cat_sum['Count']
    cr = r + len(cat_sum) + 2
    for cat, crow in cat_sum.iterrows():
        bm = sp[sp['Category']==cat].groupby('Month')['Amount'].sum()
        biggest = bm.idxmax() if not bm.empty else '-'
        bg = CAT_COLORS.get(cat, 'FFFFFFFF')
        for col, val in zip('BCDEFG', [cat, crow['Total'], crow['Pct'], int(crow['Count']), crow['Avg'], biggest]):
            c = ws[f'{col}{r}']; c.value = val; c.font = bf(10,bold=(col=='B'))
            c.fill = fill(bg); c.alignment = rgt() if col in 'CDEF' else lft(); c.border = bdr()
        ws[f'C{r}'].number_format='₹#,##0'; ws[f'D{r}'].number_format='0.0%'; ws[f'F{r}'].number_format='₹#,##0'; r+=1
    for col in 'BCDEFG':
        ws[f'{col}{r}'].font=hf(10,color='FF000000'); ws[f'{col}{r}'].fill=fill(GOLD); ws[f'{col}{r}'].border=bdr()
    ws[f'B{r}']='TOTAL'; ws[f'B{r}'].alignment=lft()
    ws[f'C{r}']=cat_sum['Total'].sum(); ws[f'C{r}'].number_format='₹#,##0'
    ws[f'D{r}']=1.0; ws[f'D{r}'].number_format='0.0%'
    ws[f'E{r}']=int(cat_sum['Count'].sum())
    r+=2; cr_start=r
    ws[f'B{r}']='Category'; ws[f'C{r}']='Amount'; r+=1
    for cat,crow in cat_sum.iterrows():
        ws[f'B{r}']=cat; ws[f'C{r}']=crow['Total']; r+=1
    pie=PieChart(); pie.title="Spending by Category"; pie.style=10
    labels=Reference(ws,min_col=2,min_row=cr_start+1,max_row=r-1)
    data=Reference(ws,min_col=3,min_row=cr_start,max_row=r-1)
    pie.add_data(data,titles_from_data=True); pie.set_categories(labels)
    pie.width=18; pie.height=14; ws.add_chart(pie,f'B{cr_start}')

def build_trends(wb, sp, months_order):
    ws = wb.create_sheet("📅 Monthly Trends")
    ws.sheet_view.showGridLines = False
    for col,w in zip('ABCDEFGH',[2,25,18,18,18,18,18,3]):
        ws.column_dimensions[col].width=w
    ws.merge_cells('B1:G2')
    ws['B1']='📅  MONTHLY SPENDING TRENDS'
    ws['B1'].font=Font(name='Arial',size=14,bold=True,color=WHITE)
    ws['B1'].fill=fill(ORANGE); ws['B1'].alignment=ctr()
    r=4; show_months=months_order[:4]
    mt=sp.groupby('Month')['Amount'].sum().reindex(show_months,fill_value=0)
    mc=sp.groupby('Month')['Amount'].count().reindex(show_months,fill_value=0)
    ma=sp.groupby('Month')['Amount'].mean().reindex(show_months,fill_value=0)
    for col,h in zip('BCDEFG',['Metric']+show_months+['']):
        c=ws[f'{col}{r}']; c.value=h; c.font=hf(10); c.fill=fill(NAVY); c.alignment=ctr(); c.border=bdr()
    r+=1
    for label,series,fmt,bg in [('Total Spend',mt,'₹#,##0',LGRAY),('Transactions',mc,'#,##0',WHITE),
                                  ('Avg Transaction',ma,'₹#,##0',LGRAY)]:
        ws[f'B{r}']=label; ws[f'B{r}'].font=bf(10,bold=True); ws[f'B{r}'].fill=fill(bg)
        ws[f'B{r}'].alignment=lft(); ws[f'B{r}'].border=bdr()
        for col,m in zip('CDEF',show_months):
            c=ws[f'{col}{r}']; c.value=series.get(m,0); c.font=bf(10)
            c.fill=fill(bg); c.alignment=rgt(); c.border=bdr(); c.number_format=fmt
        r+=1
    r+=1
    ws.merge_cells(f'B{r}:G{r}')
    ws[f'B{r}']='🌙 Weekday vs Weekend Spending'
    ws[f'B{r}'].font=Font(name='Arial',size=11,bold=True,color=WHITE)
    ws[f'B{r}'].fill=fill(PURPLE); ws[f'B{r}'].alignment=lft(); r+=1
    for col,h in zip('BCD',['Type','Total Spend','% of Total']):
        c=ws[f'{col}{r}']; c.value=h; c.font=hf(10); c.fill=fill(NAVY); c.alignment=ctr(); c.border=bdr()
    r+=1
    wday=sp.groupby('IsWeekend')['Amount'].sum(); total_wk=wday.sum()
    for is_wknd,label in [(False,'Weekday (Mon–Fri)'),(True,'Weekend (Sat–Sun)')]:
        v=wday.get(is_wknd,0)
        for col,val in zip('BCD',[label,v,v/total_wk if total_wk>0 else 0]):
            c=ws[f'{col}{r}']; c.value=val; c.font=bf(10); c.fill=fill(LGRAY)
            c.alignment=rgt() if col in 'CD' else lft(); c.border=bdr()
        ws[f'C{r}'].number_format='₹#,##0'; ws[f'D{r}'].number_format='0.0%'; r+=1
    r+=2; cr=r
    ws[f'B{r}']='Month'; ws[f'C{r}']='Total'; r+=1
    for m in show_months:
        ws[f'B{r}']=m; ws[f'C{r}']=mt.get(m,0); r+=1
    bar=BarChart(); bar.title="Monthly Total Spend"; bar.y_axis.title="Amount (₹)"
    bar.style=10; bar.type="col"
    dr=Reference(ws,min_col=3,min_row=cr,max_row=cr+len(show_months))
    cr_ref=Reference(ws,min_col=2,min_row=cr+1,max_row=cr+len(show_months))
    bar.add_data(dr,titles_from_data=True); bar.set_categories(cr_ref)
    bar.width=18; bar.height=12; ws.add_chart(bar,f'E{cr}')

def build_merchants(wb, sp):
    ws=wb.create_sheet("🏪 Top Merchants")
    ws.sheet_view.showGridLines=False
    for col,w in zip('ABCDEFG',[2,50,18,12,18,22,3]):
        ws.column_dimensions[col].width=w
    ws.merge_cells('B1:F2')
    ws['B1']='🏪  TOP MERCHANTS BY SPEND'
    ws['B1'].font=Font(name='Arial',size=14,bold=True,color=WHITE)
    ws['B1'].fill=fill(ORANGE); ws['B1'].alignment=ctr()
    r=4
    for col,h in zip('BCDEF',['Merchant','Total Spend (₹)','Times','Avg Ticket (₹)','Category']):
        c=ws[f'{col}{r}']; c.value=h; c.font=hf(10); c.fill=fill(NAVY); c.alignment=ctr(); c.border=bdr()
    r+=1
    top_m=sp.groupby('Description').agg(Total=('Amount','sum'),Count=('Amount','count')).sort_values('Total',ascending=False).head(20)
    top_m['Avg']=top_m['Total']/top_m['Count']
    top_m['Cat']=[sp[sp['Description']==d]['Category'].iloc[0] for d in top_m.index]
    medal={0:'🥇',1:'🥈',2:'🥉'}
    for i,(desc,mrow) in enumerate(top_m.iterrows()):
        bg=['FFFFF9C4','FFF3E5F5','FFE8F5E9'][i] if i<3 else (LGRAY if i%2==0 else WHITE)
        prefix=medal.get(i,f'#{i+1:2d}')
        for col,val in zip('BCDEF',[f"{prefix}  {desc}",mrow['Total'],int(mrow['Count']),mrow['Avg'],mrow['Cat']]):
            c=ws[f'{col}{r}']; c.value=val; c.font=bf(10,bold=(i<3))
            c.fill=fill(bg); c.alignment=rgt() if col in 'CDE' else lft(); c.border=bdr()
        ws[f'C{r}'].number_format='₹#,##0'; ws[f'E{r}'].number_format='₹#,##0'; r+=1
    r+=2
    ws.merge_cells(f'B{r}:F{r}')
    ws[f'B{r}']='🔄  RECURRING PAYMENTS (appears 2+ times)'
    ws[f'B{r}'].font=Font(name='Arial',size=12,bold=True,color=WHITE)
    ws[f'B{r}'].fill=fill(TEAL); ws[f'B{r}'].alignment=lft(); r+=1
    for col,h in zip('BCDE',['Merchant','Frequency','Total Paid','Category']):
        c=ws[f'{col}{r}']; c.value=h; c.font=hf(10); c.fill=fill(NAVY); c.alignment=ctr(); c.border=bdr()
    r+=1
    recur=sp.groupby('Description').filter(lambda x:len(x)>=2).groupby('Description').agg(
        Count=('Amount','count'),Total=('Amount','sum')).sort_values('Count',ascending=False).head(12)
    if not recur.empty:
        recur['Cat']=[sp[sp['Description']==d]['Category'].iloc[0] for d in recur.index]
        for desc,rrow in recur.iterrows():
            for col,val in zip('BCDE',[desc,f"{int(rrow['Count'])}x",rrow['Total'],rrow['Cat']]):
                c=ws[f'{col}{r}']; c.value=val; c.font=bf(10)
                c.fill=fill(LGRAY); c.alignment=rgt() if col=='C' else lft(); c.border=bdr()
            ws[f'C{r}'].number_format='₹#,##0'; r+=1

def build_insights(wb, sp):
    ws=wb.create_sheet("💡 Insights & Budget")
    ws.sheet_view.showGridLines=False
    for col,w in zip('ABCDEFG',[2,45,18,18,18,18,3]):
        ws.column_dimensions[col].width=w
    ws.merge_cells('B1:F2')
    ws['B1']='💡  BEHAVIOR INSIGHTS & BUDGET PLAN'
    ws['B1'].font=Font(name='Arial',size=14,bold=True,color=WHITE)
    ws['B1'].fill=fill(NAVY); ws['B1'].alignment=ctr()

    total=sp['Amount'].sum(); n_months=max(sp['Month'].nunique(),1)
    top3=sp.groupby('Category')['Amount'].sum().sort_values(ascending=False).head(3)
    misc=sp[sp['Category']=='Miscellaneous'].sort_values('Amount',ascending=False).head(5)
    high_freq=sp.groupby('Description').filter(lambda x:len(x)>=3)
    if not high_freq.empty:
        hf_top=high_freq.groupby('Description')['Amount'].sum().sort_values(ascending=False).head(3)

    overspend=[]
    for cat,amt in top3.items():
        pct=amt/total*100
        overspend.append(f"{cat}: {inr(amt)} ({pct:.1f}% of spend) — review if aligned with your priorities")

    misc_items=[f"  • {row['Description'][:45]} — {inr(row['Amount'])}" for _,row in misc.iterrows()]
    misc_note="Review these with your analyst to classify correctly next month:" if misc_items else ["None — excellent categorisation!"]

    sections=[
        ("🚨 BIGGEST SPEND AREAS", RED, overspend),
        ("🔎 UNCATEGORISED (Miscellaneous) — REVIEW THESE", ORANGE,
         [misc_note] + misc_items),
        ("💰 SAVINGS OPPORTUNITIES", TEAL, [
            "Audit all recurring subscriptions — cancel any you haven't used in 30 days",
            "Batch grocery shopping weekly instead of daily convenience deliveries (saves 15-20%)",
            "For travel: pre-book flights 3-4 weeks ahead rather than last-minute; avoid booking on EMI",
            "Set a monthly shopping cap and use a 48-hour wish-list rule before buying",
            "Review EMIs running in parallel — total EMI outflow should be <30% of income",
        ]),
        ("✅ HEALTHY HABITS TO MAINTAIN", GREEN, [
            "Keep tracking every month — consistency is the biggest financial superpower",
            "Investments via SIP/NACH are non-negotiable — protect these from lifestyle creep",
            "If weekend spend < 30% of total, you have good impulse control",
        ]),
    ]

    r=4
    for title,color,points in sections:
        ws.merge_cells(f'B{r}:F{r}')
        ws[f'B{r}']=title; ws[f'B{r}'].font=Font(name='Arial',size=11,bold=True,color=WHITE)
        ws[f'B{r}'].fill=fill(color); ws[f'B{r}'].alignment=lft(); r+=1
        for pt in points:
            ws.merge_cells(f'B{r}:F{r}')
            ws[f'B{r}']=f"   {'•' if not pt.startswith('  •') else ''}  {pt.strip()}"
            ws[f'B{r}'].font=bf(10); ws[f'B{r}'].fill=fill(LGRAY if r%2==0 else WHITE)
            ws[f'B{r}'].alignment=Alignment(horizontal='left',vertical='center',wrap_text=True)
            ws[f'B{r}'].border=bdr(); ws.row_dimensions[r].height=28; r+=1
        r+=1

    r+=1
    ws.merge_cells(f'B{r}:F{r}')
    ws[f'B{r}']='📋  SUGGESTED MONTHLY BUDGET'
    ws[f'B{r}'].font=Font(name='Arial',size=12,bold=True,color=WHITE)
    ws[f'B{r}'].fill=fill(NAVY); ws[f'B{r}'].alignment=lft(); r+=1
    for col,h in zip('BCDEF',['Category','Avg Actual/Month','Suggested Budget','Potential Saving','Tip']):
        c=ws[f'{col}{r}']; c.value=h; c.font=hf(10); c.fill=fill(NAVY); c.alignment=ctr(); c.border=bdr()
    r+=1
    # Budget = 80% of actual for overspent categories, same for healthy ones
    cat_avgs=sp.groupby('Category')['Amount'].sum()/n_months
    ta=tb=0
    for cat in sorted(cat_avgs.index):
        actual=cat_avgs[cat]; pct=actual/total*100
        budget=actual*0.8 if pct>15 else actual*0.9 if pct>8 else actual
        saving=actual-budget
        tip='Reduce by 20%' if pct>15 else 'Reduce by 10%' if pct>8 else 'Keep steady'
        bg='FFE8F5E9' if saving>100 else 'FFFFEBEE' if saving<-100 else LGRAY
        for col,val in zip('BCDEF',[cat,actual,budget,saving,tip]):
            c=ws[f'{col}{r}']; c.value=round(val) if isinstance(val,(int,float)) else val
            c.font=bf(10); c.fill=fill(bg); c.alignment=rgt() if col in 'CDE' else lft(); c.border=bdr()
        for col in 'CDE': ws[f'{col}{r}'].number_format='₹#,##0'
        ta+=actual; tb+=budget; r+=1
    for col,val in zip('BCDE',['TOTAL',ta,tb,ta-tb]):
        c=ws[f'{col}{r}']; c.value=round(val) if isinstance(val,(int,float)) else val; c.font=hf(10,color='FF000000')
        c.fill=fill(GOLD); c.alignment=rgt() if col!='B' else lft(); c.border=bdr()
        if col in 'CDE': ws[f'{col}{r}'].number_format='₹#,##0'
    ws[f'F{r}']=f"Potential saving: {inr(ta-tb)}/month"
    ws[f'F{r}'].font=Font(name='Arial',size=10,bold=True,color=GREEN)
    ws[f'F{r}'].fill=fill(GOLD); ws[f'F{r}'].border=bdr()

# ── Main ────────────────────────────────────────────────────────────────────
def main():
    parser=argparse.ArgumentParser()
    parser.add_argument('--input', required=True)
    parser.add_argument('--output', required=True)
    parser.add_argument('--name', default='Your')
    args=parser.parse_args()

    df=pd.read_csv(args.input, parse_dates=['Date'])
    df['Category']=df['Description'].apply(categorize)
    df['Month']=df['Date'].dt.strftime('%b %Y')
    df['IsWeekend']=df['Date'].dt.dayofweek>=5
    months_order=sorted(df['Month'].unique(), key=lambda m: pd.to_datetime(m, format='%b %Y'))
    sp=df[~df['Category'].isin(NON_SPEND)].copy().reset_index(drop=True)

    wb=Workbook()
    build_dashboard(wb, sp, months_order, args.name)
    build_transactions(wb, df)
    build_category(wb, sp)
    build_trends(wb, sp, months_order)
    build_merchants(wb, sp)
    build_insights(wb, sp)
    wb.save(args.output)

    print(f"✅ Report saved: {args.output}")
    print(f"   📊 {len(df)} transactions | {inr(sp['Amount'].sum())} spend | {len(months_order)} month(s)")
    misc_count=len(sp[sp['Category']=='Miscellaneous'])
    if misc_count:
        print(f"   ⚠️  {misc_count} transactions in Miscellaneous — review in Insights sheet")

if __name__=='__main__':
    main()
